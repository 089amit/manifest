# tracking.py - Export to Excel with auto-filters, leading zeros preserved, and clickable links in Service column
import os
import uuid
import re
import pandas as pd
from flask import Blueprint, request, jsonify, send_file
from utils import extract_mawb_columns
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

tracking_bp = Blueprint('tracking_mod', __name__)

TEMP_FOLDER = 'temp_files'

COLUMNS_NEEDED = {
    'HAWB': 'hawb',
    'Consignee Name': 'consignee',
    'Country': 'country',
    'NO OF PCS': 'pcs',
    'Service': 'service',
    'Forwarding Number': 'forwarding_number'
}

def keep_original_tracking_number(value):
    """Keep exact value as string, preserve leading zeros."""
    if pd.isna(value):
        return 'N/A'
    s = str(value).strip()
    return s if s != '' else 'N/A'

def extract_reference_number(df, filename):
    """Extract reference from MAWB No. column or fallback to filename."""
    possible_cols = ['MAWB No.', 'Reference', 'Shipment No.', 'MAWB Number', 'AWB No.']
    for col in possible_cols:
        if col in df.columns:
            first_val = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
            if first_val and str(first_val).strip():
                return str(first_val).strip()
    base = os.path.splitext(os.path.basename(filename))[0]
    return re.sub(r'[^a-zA-Z0-9\-_]', '_', base)

def group_tracking_data(df):
    """Group data for frontend display (unchanged)."""
    df['service'] = df['service'].fillna('Standard').replace('', 'Standard')
    df['forwarding_number'] = df['forwarding_number'].apply(keep_original_tracking_number)
    grouped = []
    for country in sorted(df['country'].unique()):
        country_df = df[df['country'] == country]
        services = []
        for service in sorted(country_df['service'].unique()):
            service_df = country_df[country_df['service'] == service]
            rows = service_df[['hawb', 'consignee', 'forwarding_number', 'pcs']].to_dict(orient='records')
            for r in rows:
                try:
                    r['pcs'] = int(r['pcs'])
                except:
                    r['pcs'] = r['pcs']
            services.append({
                'service': service,
                'count': len(service_df),
                'rows': rows
            })
        grouped.append({
            'country': country,
            'services': services
        })
    return grouped

def is_url(text):
    """Check if text is a valid URL starting with http:// or https://"""
    if not text or not isinstance(text, str):
        return False
    text = text.strip().lower()
    return text.startswith('http://') or text.startswith('https://')

def simplify_service_for_display(service):
    """Return a friendly display name for URLs."""
    if is_url(service):
        # Extract domain or return a generic label
        try:
            from urllib.parse import urlparse
            parsed = urlparse(service)
            domain = parsed.netloc.replace('www.', '')
            return f"Track on {domain}"
        except:
            return "Tracking Link"
    return service

@tracking_bp.route('/tracking/upload', methods=['POST'])
def tracking_upload():
    mawb_file = request.files.get('mawb')
    if not mawb_file:
        return jsonify({'error': 'Please upload the MAWB file.'}), 400

    session_id = str(uuid.uuid4())
    session_dir = os.path.join(TEMP_FOLDER, f'tracking_{session_id}')
    os.makedirs(session_dir, exist_ok=True)

    filepath = os.path.join(session_dir, mawb_file.filename)
    mawb_file.save(filepath)

    try:
        df = extract_mawb_columns(filepath, COLUMNS_NEEDED)
        df['forwarding_number'] = df['forwarding_number'].apply(keep_original_tracking_number)
        reference = extract_reference_number(df, mawb_file.filename)
        with open(os.path.join(session_dir, 'reference.txt'), 'w') as f:
            f.write(reference)
        # Save original service values (including URLs) for later use in Excel
        df.to_csv(os.path.join(session_dir, 'tracking_data.csv'), index=False)
        grouped = group_tracking_data(df)
        return jsonify({'session_id': session_id, 'grouped_data': grouped, 'reference': reference})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@tracking_bp.route('/tracking/data/<session_id>', methods=['GET'])
def tracking_data(session_id):
    session_dir = os.path.join(TEMP_FOLDER, f'tracking_{session_id}')
    csv_path = os.path.join(session_dir, 'tracking_data.csv')
    if not os.path.exists(csv_path):
        return jsonify({'error': 'Session not found'}), 404
    df = pd.read_csv(csv_path)
    grouped = group_tracking_data(df)
    return jsonify({'grouped_data': grouped})

@tracking_bp.route('/tracking/download/<session_id>', methods=['GET'])
def tracking_download(session_id):
    """Generate Excel file with auto-filters, column width, text formatting, and clickable links in Service column."""
    session_dir = os.path.join(TEMP_FOLDER, f'tracking_{session_id}')
    csv_path = os.path.join(session_dir, 'tracking_data.csv')
    ref_path = os.path.join(session_dir, 'reference.txt')
    if not os.path.exists(csv_path):
        return "Session not found", 404

    # Read reference
    reference = "report"
    if os.path.exists(ref_path):
        with open(ref_path, 'r') as f:
            reference = f.read().strip()
    filename = f"tracking_report_{reference}.xlsx"

    # Load data (includes original service values)
    df = pd.read_csv(csv_path)
    
    flat = []
    for _, row in df.iterrows():
        original_service = row['service'] if pd.notna(row['service']) and row['service'] != '' else 'Standard'
        flat.append({
            'Tracking Number': row['forwarding_number'],
            'HAWB': row['hawb'],
            'Consignee Name': row['consignee'],
            'Country': row['country'],
            'PCS': row['pcs'],
            'Status': '',
            'Service': original_service  # Keep original value (may be URL)
        })
    flat_df = pd.DataFrame(flat)
    column_order = ['Tracking Number', 'HAWB', 'Consignee Name', 'Country', 'PCS', 'Status', 'Service']
    flat_df = flat_df[column_order]

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tracking Report"

    # Write headers
    headers = list(flat_df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data rows with hyperlink support for Service column
    for row_idx, row in enumerate(flat_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Handle Tracking Number column (col_idx 1) - preserve leading zeros
            if col_idx == 1 and value != 'N/A':
                cell.value = str(value)
                cell.number_format = '@'
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            # Handle Service column (col_idx 7) - check if it's a URL
            elif col_idx == 7 and is_url(value):
                # Create clickable hyperlink
                display_text = simplify_service_for_display(value)
                cell.value = display_text
                cell.hyperlink = value
                cell.font = Font(color="0000FF", underline="single")
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            else:
                cell.value = value
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Auto-fit column widths
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Enable auto-filter on all columns
    ws.auto_filter.ref = ws.dimensions

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@tracking_bp.route('/tracking/download_selected', methods=['POST'])
def tracking_download_selected():
    """Generate Excel for selected rows only, with clickable links in Service column."""
    data = request.get_json()
    selected_rows = data.get('selected_rows', [])
    session_id = data.get('session_id')
    if not selected_rows:
        return jsonify({'error': 'No rows selected'}), 400

    # Build DataFrame from selected rows (these already have original service values)
    df = pd.DataFrame(selected_rows)
    df.rename(columns={
        'tracking_number': 'Tracking Number',
        'hawb': 'HAWB',
        'consignee_name': 'Consignee Name',
        'country': 'Country',
        'pcs': 'PCS',
        'status': 'Status',
        'service': 'Service'
    }, inplace=True)
    column_order = ['Tracking Number', 'HAWB', 'Consignee Name', 'Country', 'PCS', 'Status', 'Service']
    df = df[column_order]

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Selected Tracking"

    # Headers
    for col_idx, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data rows with hyperlink support for Service column
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Tracking Number column (col_idx 1)
            if col_idx == 1 and value != 'N/A':
                cell.value = str(value)
                cell.number_format = '@'
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            # Service column (col_idx 7)
            elif col_idx == 7 and is_url(value):
                display_text = simplify_service_for_display(value)
                cell.value = display_text
                cell.hyperlink = value
                cell.font = Font(color="0000FF", underline="single")
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            else:
                cell.value = value
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Auto-fit columns
    for col_idx in range(1, len(df.columns) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in range(1, len(df) + 2):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Get reference for filename
    session_dir = os.path.join(TEMP_FOLDER, f'tracking_{session_id}')
    ref_path = os.path.join(session_dir, 'reference.txt')
    reference = "report"
    if os.path.exists(ref_path):
        with open(ref_path, 'r') as f:
            reference = f.read().strip()
    filename = f"selected_tracking_{reference}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )